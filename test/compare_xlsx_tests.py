# WTFPL licence (see http://www.wtfpl.net/)
# formatted with black, see https://github.com/python/black. Usage:
#   black {source_file_or_directory} # from a command shell

import unittest
from openpyxl import Workbook
from src import compare_xlsx


class CompareXlxsTestCase(unittest.TestCase):
    def test_compare_lines_works_with_OK(self):
        actual_book = Workbook()
        actual_sheet = actual_book.active
        actual_sheet["A2"] = "Test"
        actual_sheet["B2"] = 1
        actual_sheet["C2"] = None
        expected_book = Workbook()
        expected_sheet = expected_book.active
        expected_sheet["A2"] = "Test"
        expected_sheet["B2"] = 1
        expected_sheet["C2"] = None
        actual_book, expected_book, error_detected = compare_xlsx.compare_lines(
            actual_book, expected_book, "C"
        )
        for i in range(1, 3):
            actual_sheet = actual_book.active
            actual_cell = actual_sheet.cell(row=2, column=i)
            actual_color = actual_cell.fill.start_color.index
            self.assertEqual("0000FF00", actual_color)
            expected_sheet = expected_book.active
            actual_cell = expected_sheet.cell(row=2, column=i)
            actual_color = actual_cell.fill.start_color.index
            self.assertEqual("0000FF00", actual_color)

    def test_compare_lines_works_with_KO(self):
        actual_book = Workbook()
        actual_sheet = actual_book.active
        actual_sheet["A2"] = "Test"
        actual_sheet["B2"] = 1
        actual_sheet["C2"] = None
        expected_book = Workbook()
        expected_sheet = expected_book.active
        expected_sheet["A2"] = "Something else"
        expected_sheet["B2"] = 1
        expected_sheet["C2"] = None
        actual_book, expected_book, error_detected = compare_xlsx.compare_lines(
            actual_book, expected_book, "C"
        )
        actual_sheet = actual_book.active
        actual_color = actual_sheet["A2"].fill.start_color.index
        self.assertEqual("00FF0000", actual_color)
        expected_sheet = expected_book.active
        actual_color = expected_sheet["A2"].fill.start_color.index
        self.assertEqual("00FF0000", actual_color)
        for i in range(2, 3):
            actual_sheet = actual_book.active
            actual_cell = actual_sheet.cell(row=2, column=i)
            actual_color = actual_cell.fill.start_color.index
            self.assertEqual("0000FF00", actual_color)
            expected_sheet = expected_book.active
            actual_cell = expected_sheet.cell(row=2, column=i)
            actual_color = actual_cell.fill.start_color.index
            self.assertEqual("0000FF00", actual_color)


if __name__ == "__main__":
    unittest.main()
