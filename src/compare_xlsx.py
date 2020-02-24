# WTFPL licence (see http://www.wtfpl.net/)
# formatted with black, see https://github.com/python/black.


import argparse
from openpyxl import utils, Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import RED, GREEN
from os import path
import csv


def compare_lines(actual_book, expected_book, max_range):
    actual_sheet = actual_book.active
    expected_sheet = expected_book.active
    max_width = utils.cell.column_index_from_string(max_range)
    max_depth = max(actual_sheet.max_row, expected_sheet.max_row)
    error_detected = False
    for i in range(2, max_depth + 1):
        for j in range(1, max_width + 1):
            actual_cell = actual_sheet.cell(row=i, column=j)
            expected_cell = expected_sheet.cell(row=i, column=j)
            if actual_cell.value == expected_cell.value:
                actual_cell.fill = PatternFill(fgColor=GREEN, fill_type="solid")
                expected_cell.fill = PatternFill(fgColor=GREEN, fill_type="solid")
            else:
                actual_cell.fill = PatternFill(fgColor=RED, fill_type="solid")
                expected_cell.fill = PatternFill(fgColor=RED, fill_type="solid")
                error_detected = True
    return actual_book, expected_book, error_detected


def compare_xlsx_sheets(actual_file, expected_file, max_range):
    actual_book, actual_xlsx = convert_csv_to_file(actual_file)
    expected_book, expected_xlsx = convert_csv_to_file(expected_file)
    actual_book, expected_book, error_detected = compare_lines(
        actual_book, expected_book, max_range
    )
    actual_book.save(actual_xlsx)
    expected_book.save(expected_xlsx)
    print("-----------------------------------")
    print("")
    if error_detected:
        print("Error detected, please check the file " + actual_xlsx)
    else:
        print("Everything is OK!")


def convert_csv_to_file(file_full_path):
    file_path, file_name = path.split(file_full_path)
    print("csv_file: " + str(file_full_path))
    xlsx_file = path.join(file_path, str(file_name.rpartition(".")[0]) + ".xlsx")
    print("xlsx_file: " + str(xlsx_file))
    book = Workbook()
    sheet = book.worksheets[0]
    sheet.title = "work_sheet"
    with open(file_full_path) as f:
        csv.register_dialect("semi-colon", delimiter=";")
        reader = csv.reader(f, dialect="semi-colon")
        for i, row in enumerate(reader):
            for j, cell in enumerate(row):
                if cell is not None:
                    cell = str(cell).strip()
                sheet.cell(row=i + 1, column=j + 1).value = cell
    return book, xlsx_file


def read_args():
    document_string = """ 
    This script reads two csv files and a maximum range letter.
    It converts the csv files to xlsx files and compares them 
    line by line, cell by cell until the indicated max range. 
    Then it modifies the xlsx files by changing the background color of
    the cells: green if they match, red if not."""
    parser = argparse.ArgumentParser(description=document_string)
    help_string = "actual csv file path and name"
    parser.add_argument("-a", "--actual", help=help_string)
    help_string = "expected csv file path and name"
    parser.add_argument("-e", "--expected", help=help_string)
    help_string = "maximum range to test, letter expected"
    parser.add_argument("-m", "--max_range", help=help_string)
    return parser.parse_args()


if __name__ == "__main__":
    args = read_args()
    compare_xlsx_sheets(args.actual, args.expected, args.max_range)
    print("")
